import time
from selenium import webdriver
import requests
from openpyxl import load_workbook
import random
import pandas as pd

wb = load_workbook('mipt_jokes_music_access.xlsx')   #открывает айдишки
ws = wb.active

login = ''       #свой логин от вк
pas = ''      #свой пароль от вк
mes = ['Добро пожаловать']

s = requests.Session()

driver = webdriver.Chrome(executable_path='/path/to/chromedriver')
driver.set_window_size(1366, 1660)
driver.implicitly_wait(3)
driver.get('https://vk.com')
#time.sleep(random.randint(3,10))
username = driver.find_element_by_id('index_email')
username.clear()
username.send_keys(login)
password = driver.find_element_by_id('index_pass')
password.clear()
password.send_keys(pas)
log = driver.find_element_by_id('index_login_button')
time.sleep(random.randint(3,5))

log.click()
driver.get('https://vk.com/audios69')
top=dict()
for i in range (1,1306):      #тут количество строк в предыдущем xlsx
    audio_id='audios'+str(ws.cell(row=i, column=3).value)
    driver.get('https://vk.com/{}'.format(audio_id))
    data = driver.find_elements_by_class_name("artist_link")
    for o in range(1,len(data)):
        if data[o].text in top:
            top[data[o].text] += 1
        else:
            top[data[o].text] = 1
        print(data[o].text)
    print(top)
    #time.sleep(180)

df = pd.DataFrame(data=top, index=[0])
df = (df.T)
print (df)
df.to_excel('test_bot_df.xlsx')
