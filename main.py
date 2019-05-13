import vk
from settings import s_token, token, N
from openpyxl import Workbook

session = vk.Session()
api = vk.API(session, v=5.92)

ids = [None]*N              #число означает количество тысяч пользователей в паблике, т к vk api умеет выгружать только по 1000 пользователей
users_raw = [None]*N        #
users_total =[]
for i in range (0, N):      #
    ids[i]=api.groups.getMembers(group_id = 'test_bot_df', offset = i*1000, access_token = token)['items']
    users_raw[i] = api.users.get(user_ids=str(ids[i]).strip('[]'), fields='can_see_audio', access_token=s_token)
    users_total = users_total+users_raw[i]


wb = Workbook()
ws = wb.active

for i in range(1, len(users_total)):
    if 'can_see_audio' in users_total[i]:
        if users_total[i]['can_see_audio'] != 0:
            ws.cell(row=i, column=1).value=users_total[i]['first_name']
            ws.cell(row=i, column=2).value = users_total[i]['last_name']
            ws.cell(row=i, column=3).value = users_total[i]['id']
wb.save('test_bot_df.xlsx')     #сохраняет айдишки

#for i in range(len(users_total)):
#    print(users_total[i]['can_see_audio'])

print(users_total)
print(len(users_total))